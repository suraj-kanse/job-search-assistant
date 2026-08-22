import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

TRACKER_PATH = os.path.join(os.path.dirname(__file__), "job_tracker.xlsx")

HEADERS = [
    "Date Found", "Company", "Role", "Job ID", "Platform", "Location", 
    "Posted Date", "Job URL", "Fitness Score", "Resume Generated", 
    "Cover Letter", "Status", "Date Applied", "Response Date", "Outcome", 
    "Interview Stage", "Rejection Reason", "Notes", "Follow-up Date", "Next Action"
]

def init_tracker():
    """Initializes the Excel tracker with header formatting if it does not exist."""
    if os.path.exists(TRACKER_PATH):
        return openpyxl.load_workbook(TRACKER_PATH)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Write headers
    ws.append(HEADERS)
    
    # Format headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.row_dimensions[1].height = 28
    
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    wb.save(TRACKER_PATH)
    return wb

def load_tracker():
    """Loads the job tracker workbook."""
    init_tracker()
    return openpyxl.load_workbook(TRACKER_PATH)

def add_jobs_to_tracker(jobs):
    """Adds a list of found jobs to the tracker if they aren't already in it."""
    wb = load_tracker()
    ws = wb["Applications"]
    
    # Get existing Job IDs to prevent duplicates
    existing_ids = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=4).value # Column 4 is Job ID
        if val:
            existing_ids.add(val)
            
    today_str = datetime.date.today().strftime("%d %b %Y")
    new_jobs_added = 0
    
    for job in jobs:
        if job["job_id"] in existing_ids:
            continue
            
        row_data = [
            today_str,                                     # Date Found
            job["company"],                                # Company
            job["title"],                                  # Role
            job["job_id"],                                 # Job ID
            job["platform"],                               # Platform
            job["location"],                               # Location
            job["posted_date"],                            # Posted Date
            job["link"],                                   # Job URL
            f"{job['fitness_score']}%",                    # Fitness Score
            "✅ Ready",                                    # Resume Generated
            "✅ Ready",                                    # Cover Letter
            "Ready to Apply",                              # Status
            "", "", "", "", "", "", "", ""                 # Applied stats (empty)
        ]
        
        ws.append(row_data)
        
        # Center align status, fitness, platform, location, dates
        center_cols = [1, 4, 5, 6, 7, 9, 10, 11, 12]
        new_row = ws.max_row
        for col in center_cols:
            ws.cell(row=new_row, column=col).alignment = Alignment(horizontal="center")
            
        new_jobs_added += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid long URL styling from messing up column width
            val_str = str(cell.value or '')
            if cell.column == 8 and len(val_str) > 30: # Job URL
                val_str = "Link"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(TRACKER_PATH)
    print(f"Added {new_jobs_added} new jobs to tracker.")
    return new_jobs_added

def update_job_status(job_id, new_status, extra_fields=None):
    """Updates the status and other fields of a specific job by Job ID."""
    wb = load_tracker()
    ws = wb["Applications"]
    
    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == job_id:
            ws.cell(row=row, column=12).value = new_status # Column 12 is Status
            
            # If status is "Applied", set Date Applied
            if new_status.lower() == "applied":
                ws.cell(row=row, column=13).value = datetime.date.today().strftime("%d %b %Y") # Date Applied
                
            if extra_fields:
                for field_name, value in extra_fields.items():
                    if field_name in HEADERS:
                        col_idx = HEADERS.index(field_name) + 1
                        ws.cell(row=row, column=col_idx).value = value
            updated = True
            break
            
    if updated:
        wb.save(TRACKER_PATH)
        print(f"Updated job {job_id} to status: {new_status}")
    else:
        print(f"Job ID {job_id} not found in tracker.")
    return updated

def get_tracker_summary():
    """Returns a dictionary of summary statistics from the tracker."""
    if not os.path.exists(TRACKER_PATH):
        return {"total": 0, "statuses": {}}
        
    wb = load_tracker()
    ws = wb["Applications"]
    
    total = ws.max_row - 1
    stats = {}
    
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=12).value or "Found"
        stats[status] = stats.get(status, 0) + 1
        
    return {
        "total": total,
        "statuses": stats
    }

if __name__ == "__main__":
    init_tracker()
    print("Tracker initialized successfully.")
    
    # Test adding a mock job
    mock_job = {
        "company": "TestCorp",
        "title": "Backend Engineer",
        "job_id": "TEST-JD-12345",
        "platform": "LinkedIn",
        "location": "Remote",
        "posted_date": "22 Aug 2026",
        "link": "https://test.com/job",
        "fitness_score": 95
    }
    
    add_jobs_to_tracker([mock_job])
    print("Summary:", get_tracker_summary())
    update_job_status("TEST-JD-12345", "Applied", {"Notes": "Applied via website"})
    print("Summary after update:", get_tracker_summary())
